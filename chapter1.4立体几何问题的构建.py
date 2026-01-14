# %% 第一章 解析方法与几何模型 - 1.4 立体几何模型构建
import numpy as np
from scipy.optimize import fsolve
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment  # 用于设置对齐方式

# 常量定义
theta = 2 * np.pi / 3  # 全开角
alpha = -1.5 / 180 * np.pi  # 坡度
htheta = theta / 2  # 半开角
h = 120  # 正中心海水深度
unit = 1852  # 一海里等于1852米
k0 = np.tan(alpha)  # 海底坡面直线斜率

# 初始化覆盖宽度矩阵
W = np.zeros((8, 8))

# 计算不同角度和位置的覆盖宽度

for i in range(1, 9):
    for j in range(1, 9):
        beta = (i - 1) * np.pi / 4
        d = (j - 1) * 0.3 * unit
        v = np.array([np.cos(beta), np.sin(beta), 0])  # 直线法向量
        origin = v * d  # 发射位置

        # 波束的方向向量
        v1 = np.array([-np.sin(beta) * np.sin(htheta), np.cos(beta) * np.sin(htheta), -np.cos(htheta)])
        v2 = np.array([np.sin(beta) * np.sin(htheta), -np.cos(beta) * np.sin(htheta), -np.cos(htheta)])
        # 用于找到波束与海底交点的函数
        leftsolve = lambda t: (v1[0] * t + origin[0]) * k0 - h - (v1[2] * t + origin[2])
        rightsolve = lambda t: (v2[0] * t + origin[0]) * k0 - h - (v2[2] * t + origin[2])

        # 解方程找到交点
        tleft = fsolve(leftsolve, 0)
        tright = fsolve(rightsolve, 0)

        # 计算左右交点的坐标
        pleft = v1 * tleft + origin
        pright = v2 * tright + origin

        # 计算并存储覆盖宽度
        W[i - 1, j - 1] = np.linalg.norm(pleft - pright)
# 打印覆盖宽度矩阵
print("覆盖宽度矩阵 W:\n", W)

# 将数据放入表格当中
# 1. 定义表格的行/列标签（与你的表格完全对应）
# 行标签：测线方向夹角（°）
row_labels = [0, 45, 90, 135, 180, 225, 270, 315]
# 列标签：测量船距中心点的距离（海里）
col_labels = [0, 0.3, 0.6, 0.9, 1.2, 1.5, 1.8, 2.1]

# 2. 初始化空表格（DataFrame）
# 表格标题：行名=“测线方向夹角/°”，列名=“测量船距海域中心点处的距离/海里”，单元格存覆盖宽度
df = pd.DataFrame(
    index=row_labels,  # 行标签（测线方向）
    columns=col_labels,  # 列标签（距离）
    dtype=float  # 单元格数据类型（覆盖宽度是数值）
)
# 设置行列的“标题名”（更贴合你的表格格式）
df.index.name = "测线方向夹角/°"
# df.columns.name = "测量船距海域中心点处的距离/海里"

# 3. 填入数字（示例：定位单元格并赋值）
# 格式：df.loc[测线方向, 距离] = 覆盖宽度
for i in range(1, 9):
    for j in range(1, 9):
        df.iloc[i - 1, j - 1] = W[i - 1, j - 1]

# 4. 查看当前表格（可以在PyCharm控制台直接打印）
print("当前覆盖宽度表格：")
print(df)

# 5. （可选）导出为Excel文件（方便后续编辑）
save_path = r"D:\数模导论研习社\chapter1.4result.xlsx"  # 自定义保存路径
with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
    # 先写一个空的DataFrame，预留标题行/列的位置
    # 步骤1：写列标题（“测量船距海域中心点处的距离/海里”）
    # 先创建一个1行的DataFrame，第一列空，后面是列标题名（合并单元格用）
    col_title = pd.DataFrame([[df.columns.name]]).reindex(columns=range(1 + len(df.columns)))
    col_title.to_excel(writer, sheet_name="覆盖宽度表", startrow=0, startcol=0, index=False, header=False)

    # 步骤2：写数据表格（包含行索引和列索引）
    # 从第1行、第0列开始写（跳过列标题行）
    df.to_excel(writer, sheet_name="覆盖宽度表", startrow=1, startcol=0)

    # 步骤3：设置行标题（“测线方向夹角/°”）
    # 获取Excel的工作表对象，手动修改单元格
    worksheet = writer.sheets["覆盖宽度表"]
    # 把“测线方向夹角/°”写入第0行、第0列（与行索引列合并）
    worksheet["A2"] = df.index.name

# 用openpyxl修改格式（合并单元格+居中+填文字）
# 加载已导出的Excel文件
wb = load_workbook(save_path)
ws = wb["覆盖宽度表"]  # 获取工作表

# 关键操作1：合并第一行2-9列（Excel列标B-I，对应代码中range(2,10)）
# 合并语法：ws.merge_cells("起始单元格:结束单元格")
ws.merge_cells("B1:I1")  # 第一行B列到I列（2-9列）合并

# 关键操作2：在合并后的单元格中填入指定文字
target_text = "测量船距海域中心点处的距离/海里"  # 你要填的文字
ws["B1"] = target_text  # 合并后的单元格只需给起始单元格赋值

# 关键操作3：设置单元格文字居中对齐（水平+垂直居中）
ws["B1"].alignment = Alignment(
    horizontal="center",  # 水平居中
    vertical="center"  # 垂直居中
)

# 步骤4：保存修改后的Excel文件
wb.save(save_path)
wb.close()  # 关闭文件，释放资源

print(f"Excel文件已生成！路径：{save_path}")
print("已完成：第一行2-9列合并单元格+居中+填入文字「测量船距海域中心点处的距离/海里」")
